from tempfile import NamedTemporaryFile
import openpyxl


def create_excel(
    keywords,
    list_taxa_id,
    all_strain_data,
    create_private_project_id,
    unique_study_id_val,
):
    """
    keywords:        NCBI Taxonomy Ids of taxa with exact match to a NCBI Taxonomy Id
    list_taxa_id:    names of the taxa with exact NCBI Taxonomy ids
    all_strain_data: list of dictionaries with taxa without an exact NCBI Taxonomy Id
    """

    wb = openpyxl.load_workbook('templates/excel/metadata_template.xlsx')
    sheet = wb['COMMUNITY_MEMBERS']

    k = 0  # Initialize k here

    if keywords:

        for k in range(1, len(keywords) + 1):
            sheet.cell(row=k+1, column=1, value=f'member_{k}')
            sheet.cell(row=k+1, column=2, value=1)
            sheet.cell(row=k+1, column=3, value=keywords[k-1])
            sheet.cell(row=k+1, column=4, value=list_taxa_id[k-1])

    if all_strain_data:

        for i, dic in enumerate(all_strain_data):

            index = dic["case_number"]
            name = dic.get('name', '')

            if name is None or name == "":
                continue

            description = dic.get('description', '')
            parent_ncbi_tax_id = dic.get('parent_taxon_id')

            sheet.cell(row=k+i+2, column=1, value=f'member_{k+i+1}')
            sheet.cell(row=k+i+2, column=2, value=0)
            sheet.cell(row=k+i+2, column=3, value=name)
            sheet.cell(row=k+i+2, column=4, value=parent_ncbi_tax_id)
            # TODO (2025-03-12) Assembly ID in "new strain" form:
            sheet.cell(row=k+i+2, column=5, value='')
            sheet.cell(row=k+i+2, column=6, value=description)

    sheet_study = wb['STUDY']
    if create_private_project_id is not None:
        sheet_study.cell(row=2, column=1, value=create_private_project_id)
    if unique_study_id_val is not None:
        sheet_study.cell(row=2, column=2, value=unique_study_id_val)

    temp_file = NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()  # Close the file immediately

    # Read the contents of the temporary file in binary mode
    with open(temp_file.name, 'rb') as f:
        data_excel = f.read()

    return data_excel
